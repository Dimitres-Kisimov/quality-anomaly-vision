"""Deliverables build end-to-end on a tiny config (torch required, else skip)."""

import pytest

torch = pytest.importorskip("torch")

import openpyxl  # noqa: E402

from qav.data import DataConfig  # noqa: E402
from qav.evaluate import run_full_evaluation  # noqa: E402
from qav.exports import build_deliverables  # noqa: E402
from qav.model import TrainConfig  # noqa: E402


def test_deliverables_and_figures_are_written_and_nonempty(tmp_path):
    report = run_full_evaluation(
        DataConfig(n_train=40, n_test_clean=12, n_test_defective=12, seed=13),
        TrainConfig(epochs=2, batch_size=16, seed=3),
    )
    assert len(report.methods) == 3
    assert report.recommendation.method in {m.name for m in report.methods}

    out = tmp_path / "deliverables"
    figs = tmp_path / "figures"
    sizes = build_deliverables(report, out, figs)

    pdf = out / "qa_defect_report.pdf"
    xlsx = out / "qa_defect_metrics.xlsx"
    assert pdf.exists() and sizes[str(pdf)] > 10_240, "PDF should be a real multi-page report"
    assert xlsx.exists() and sizes[str(xlsx)] > 4_096

    wb = openpyxl.load_workbook(xlsx)
    assert set(wb.sheetnames) == {
        "Metrics", "PerDefectType", "Economics", "Robustness", "Assumptions", "PerImageScores"
    }
    metrics = wb["Metrics"]
    assert metrics.max_row == 4, "header + one row per method"
    assert wb["PerImageScores"].max_row == 25, "header + one row per test image"
    econ = wb["Economics"]
    assert econ.max_row >= 4, "header + one row per operating point"
    assert econ["A1"].value == "threshold" and econ["M1"].value == "is_recommended"
    # Exactly one operating point is flagged as recommended.
    flags = [econ.cell(row=r, column=13).value for r in range(2, econ.max_row + 1)]
    assert sum(int(v) for v in flags) == 1

    # Robustness sheet: header, one baseline row per method, then the sweep grid.
    rob = wb["Robustness"]
    assert rob["A1"].value == "method" and rob["I1"].value == "d_tpr_at_5pct_fpr"
    assert rob.max_row == 1 + 3 + 3 * 12, "header + 3 baselines + 3 methods x 12 sweep points"
    baseline_rows = [
        rob.cell(row=r, column=2).value for r in range(2, rob.max_row + 1)
    ].count("none")
    assert baseline_rows == 3, "one clean-baseline row per method"

    # The byte-identical cost-curve + robustness deliverables the README references.
    csv_file = out / "cost_curve.csv"
    svg_file = figs / "cost_curve.svg"
    robustness_csv = out / "robustness.csv"
    assert csv_file.exists() and sizes[str(csv_file)] > 0
    assert svg_file.exists() and sizes[str(svg_file)] > 0
    assert robustness_csv.exists() and sizes[str(robustness_csv)] > 0
    assert csv_file.read_text(encoding="utf-8").startswith("threshold,reject_rate,")
    assert svg_file.read_text(encoding="utf-8").startswith("<svg")
    assert robustness_csv.read_text(encoding="utf-8").startswith("method,perturbation,severity,")

    for name in ("gallery.png", "roc_pr.png", "per_type_auc.png", "robustness.png"):
        f = figs / name
        assert f.exists() and f.stat().st_size > 1_024, f"{name} should be a real image"
